from openpyxl import load_workbook
import pandas as pd
from tkinter import *
from tkinter.messagebox import showinfo
from functools import partial
from settings import *
# from pandas.core.ops import key

class Match:
    def __init__(self, data):
        self.date = data[col_dict['date'] - 1]
        self.team1 = data[col_dict['team1'] - 1]
        self.team2 = data[col_dict['team2'] - 1]
        self.tp_h = data[col_dict['tp_h'] - 1]
        self.tp_a = data[col_dict['tp_a'] - 1]
        self.ap_h = data[col_dict['ap_h'] - 1]
        self.ap_a = data[col_dict['ap_a'] - 1]
        self.sht_h = data[col_dict['sht_h'] - 1]
        self.sht_a = data[col_dict['sht_a'] - 1]
        self.g_h = data[col_dict['g_h'] - 1]
        self.g_a = data[col_dict['g_a'] - 1]
        self.sht_all = data[col_dict['sht_all'] - 1]
        self.g_all = data[col_dict['g_all'] - 1]
        self.result = data[col_dict['result'] - 1]
        self.ft = data[col_dict['ft'] - 1]


def parse_main_data(ws, rows_number, cols_number):
    matches = []
    for row in range(2, rows_number + 2):
        data = [ws.cell(row, col).value for col in range(1, cols_number + 1)]
        matches.append(Match(data))
    return matches


def find_cell(ws, value):
    for col in range(1, 75):
        for row in range(1, 47):
            if ws.cell(row, col).value == value:
                return row, col


def parse_team_names(ws, key_word):
    start_row, col = find_cell(ws, key_word)
    list_matches = []
    for row in range(start_row + 1, start_row + 30):
        list_matches.append(ws.cell(row, col).value)
        if ws.cell(row + 1, col).value is None:
            return list_matches


def parse_last_matches(ws, key_word):
    start_row, start_col = find_cell(ws, key_word)
    list_matches = []
    for row in range(start_row + 1, start_row + 15):
        list_matches.append((ws.cell(row, start_col).value, ws.cell(row, start_col + 1).value))
        if ws.cell(row + 1, start_col).value is None:
            return list_matches


def count_for(matches, property, team, last_matches_number):
    counter = 0
    i = last_matches_number
    for match in reversed(matches):
        if match.team1 == team:
            property1 = property + '_h'
            expr = 'match.' + property1
            if i:
                counter += eval(expr)
                i -= 1
            else:
                break
        if match.team2 == team:
            property2 = property + '_a'
            expr = 'match.' + property2
            if i:
                counter += eval(expr)
                i -= 1
            else:
                break
    return counter


def count_against(matches, property, team, last_matches_number):
    counter = 0
    i = last_matches_number
    for match in reversed(matches):
        if match.team1 == team:
            property1 = property + '_a'
            expr = 'match.' + property1
            if i:
                counter += eval(expr)
                i -= 1
            else:
                break
        if match.team2 == team:
            property2 = property + '_h'
            expr = 'match.' + property2
            if i:
                counter += eval(expr)
                i -= 1
            else:
                break
    return counter


def count_over_under_all(matches, param, property, team, last_matches_report2):
    over = 0
    under = 0
    i = last_matches_report2
    for match in reversed(matches):
        expr = 'match.' + property
        if match.team1 == team or match.team2 == team:
            if i:
                if eval(expr) > param:
                    over += 1
                if eval(expr) < param:
                    under += 1
                i -= 1
            else:
                break
    return over, under


def count_over_under_offensive(matches, param, property, team, last_matches_report2):
    over = 0
    under = 0
    i = last_matches_report2
    for match in reversed(matches):
        if match.team1 == team:
            property1 = property + '_h'
            expr = 'match.' + property1
            if i:
                if eval(expr) > param:
                    over += 1
                if eval(expr) < param:
                    under += 1
                i -= 1
            else:
                break
        if match.team2 == team:
            property2 = property + '_a'
            expr = 'match.' + property2
            if i:
                if eval(expr) > param:
                    over += 1
                if eval(expr) < param:
                    under += 1
                i -= 1
            else:
                break
    return over, under


def count_over_under_defensive(matches, param, property, team, last_matches_report2):
    over = 0
    under = 0
    i = last_matches_report2
    for match in reversed(matches):
        if match.team1 == team:
            property1 = property + '_a'
            expr = 'match.' + property1
            if i:
                if eval(expr) > param:
                    over += 1
                if eval(expr) < param:
                    under += 1
                i -= 1
            else:
                break
        if match.team2 == team:
            property2 = property + '_h'
            expr = 'match.' + property2
            if i:
                if eval(expr) > param:
                    over += 1
                if eval(expr) < param:
                    under += 1
                i -= 1
            else:
                break
    return over, under


def add_report_1(ws, matches, teams, rep1_addr, rep1_last):
    last_matches_report1 = rep1_last
    dest_row_rep1 = int(re.findall(r'\d+', rep1_addr)[0])
    # print(dest_row_rep1)
    dest_col_rep1 = re.findall(r'[a-zA-Z]+', rep1_addr)[0]
    dest_col_rep1 = cols.index(dest_col_rep1)
    # print(dest_col_rep1)
    # print(teams)
    for i in range(len(teams)):
        # колонка кол-во игр
        # print(ws.cell(dest_row_rep1 + i, dest_col_rep1 + 0).value())
        ws.cell(dest_row_rep1 + i, dest_col_rep1 + 0).value = last_matches_report1
        # for (за)
        tp = count_for(matches, 'tp', teams[i], last_matches_report1)
        ws.cell(dest_row_rep1 + i, dest_col_rep1 + 1).value = tp
        ap = count_for(matches, 'ap', teams[i], last_matches_report1)
        ws.cell(dest_row_rep1 + i, dest_col_rep1 + 2).value = ap
        sht = count_for(matches, 'sht', teams[i], last_matches_report1)
        ws.cell(dest_row_rep1 + i, dest_col_rep1 + 3).value = sht
        g = count_for(matches, 'g', teams[i], last_matches_report1)
        ws.cell(dest_row_rep1 + i, dest_col_rep1 + 4).value = g
        # against (против)
        tp = count_against(matches, 'tp', teams[i], last_matches_report1)
        ws.cell(dest_row_rep1 + i, dest_col_rep1 + 5).value = tp
        ap = count_against(matches, 'ap', teams[i], last_matches_report1)
        ws.cell(dest_row_rep1 + i, dest_col_rep1 + 6).value = ap
        sht = count_against(matches, 'sht', teams[i], last_matches_report1)
        ws.cell(dest_row_rep1 + i, dest_col_rep1 + 7).value = sht
        g = count_against(matches, 'g', teams[i], last_matches_report1)
        ws.cell(dest_row_rep1 + i, dest_col_rep1 + 8).value = g


def add_report_2(ws, matches, last_matches, rep2_addr, rep2_last):
    last_matches_report2 = rep2_last
    dest_row_rep2 = int(re.findall(r'\d+', rep2_addr)[0])
    dest_col_rep2 = re.findall(r'[a-zA-Z]+', rep2_addr)[0]
    dest_col_rep2 = cols.index(dest_col_rep2)
    i = 0
    for team1, team2 in last_matches:
        # team1, team2
        ws.cell(dest_row_rep2 + i, dest_col_rep2).value = team1
        ws.cell(dest_row_rep2 + i, dest_col_rep2 + 1).value = team2
        # sht_all
        team1_over, team1_under = count_over_under_all(matches, 6.5, 'sht_all', team1, last_matches_report2)
        team2_over, team2_under = count_over_under_all(matches, 6.5, 'sht_all', team2, last_matches_report2)
        ws.cell(dest_row_rep2 + i, dest_col_rep2 + 2).value = team1_over
        ws.cell(dest_row_rep2 + i, dest_col_rep2 + 3).value = team1_under
        ws.cell(dest_row_rep2 + i, dest_col_rep2 + 4).value = team2_over
        ws.cell(dest_row_rep2 + i, dest_col_rep2 + 5).value = team2_under
        team1_over, team1_under = count_over_under_all(matches, 9.5, 'sht_all', team1, last_matches_report2)
        team2_over, team2_under = count_over_under_all(matches, 9.5, 'sht_all', team2, last_matches_report2)
        ws.cell(dest_row_rep2 + i, dest_col_rep2 + 6).value = team1_over
        ws.cell(dest_row_rep2 + i, dest_col_rep2 + 7).value = team1_under
        ws.cell(dest_row_rep2 + i, dest_col_rep2 + 8).value = team2_over
        ws.cell(dest_row_rep2 + i, dest_col_rep2 + 9).value = team2_under
        team1_over, team1_under = count_over_under_all(matches, 12.5, 'sht_all', team1, last_matches_report2)
        team2_over, team2_under = count_over_under_all(matches, 12.5, 'sht_all', team2, last_matches_report2)
        ws.cell(dest_row_rep2 + i, dest_col_rep2 + 10).value = team1_over
        ws.cell(dest_row_rep2 + i, dest_col_rep2 + 11).value = team1_under
        ws.cell(dest_row_rep2 + i, dest_col_rep2 + 12).value = team2_over
        ws.cell(dest_row_rep2 + i, dest_col_rep2 + 13).value = team2_under
        # g_all
        team1_over, team1_under = count_over_under_all(matches, 1.5, 'g_all', team1, last_matches_report2)
        team2_over, team2_under = count_over_under_all(matches, 1.5, 'g_all', team2, last_matches_report2)
        ws.cell(dest_row_rep2 + i, dest_col_rep2 + 14).value = team1_over
        ws.cell(dest_row_rep2 + i, dest_col_rep2 + 15).value = team1_under
        ws.cell(dest_row_rep2 + i, dest_col_rep2 + 16).value = team2_over
        ws.cell(dest_row_rep2 + i, dest_col_rep2 + 17).value = team2_under
        team1_over, team1_under = count_over_under_all(matches, 2.5, 'g_all', team1, last_matches_report2)
        team2_over, team2_under = count_over_under_all(matches, 2.5, 'g_all', team2, last_matches_report2)
        ws.cell(dest_row_rep2 + i, dest_col_rep2 + 18).value = team1_over
        ws.cell(dest_row_rep2 + i, dest_col_rep2 + 19).value = team1_under
        ws.cell(dest_row_rep2 + i, dest_col_rep2 + 20).value = team2_over
        ws.cell(dest_row_rep2 + i, dest_col_rep2 + 21).value = team2_under
        team1_over, team1_under = count_over_under_all(matches, 3.5, 'g_all', team1, last_matches_report2)
        team2_over, team2_under = count_over_under_all(matches, 3.5, 'g_all', team2, last_matches_report2)
        ws.cell(dest_row_rep2 + i, dest_col_rep2 + 22).value = team1_over
        ws.cell(dest_row_rep2 + i, dest_col_rep2 + 23).value = team1_under
        ws.cell(dest_row_rep2 + i, dest_col_rep2 + 24).value = team2_over
        ws.cell(dest_row_rep2 + i, dest_col_rep2 + 25).value = team2_under
        team1_over, team1_under = count_over_under_all(matches, 4.5, 'g_all', team1, last_matches_report2)
        team2_over, team2_under = count_over_under_all(matches, 4.5, 'g_all', team2, last_matches_report2)
        ws.cell(dest_row_rep2 + i, dest_col_rep2 + 26).value = team1_over
        ws.cell(dest_row_rep2 + i, dest_col_rep2 + 27).value = team1_under
        ws.cell(dest_row_rep2 + i, dest_col_rep2 + 28).value = team2_over
        ws.cell(dest_row_rep2 + i, dest_col_rep2 + 29).value = team2_under
        # sht_offensive
        team1_over, team1_under = count_over_under_offensive(matches, 3.5, 'sht', team1, last_matches_report2)
        team2_over, team2_under = count_over_under_offensive(matches, 3.5, 'sht', team2, last_matches_report2)
        ws.cell(dest_row_rep2 + i, dest_col_rep2 + 30).value = team1_over
        ws.cell(dest_row_rep2 + i, dest_col_rep2 + 31).value = team1_under
        ws.cell(dest_row_rep2 + i, dest_col_rep2 + 32).value = team2_over
        ws.cell(dest_row_rep2 + i, dest_col_rep2 + 33).value = team2_under
        team1_over, team1_under = count_over_under_offensive(matches, 6.5, 'sht', team1, last_matches_report2)
        team2_over, team2_under = count_over_under_offensive(matches, 6.5, 'sht', team2, last_matches_report2)
        ws.cell(dest_row_rep2 + i, dest_col_rep2 + 34).value = team1_over
        ws.cell(dest_row_rep2 + i, dest_col_rep2 + 35).value = team1_under
        ws.cell(dest_row_rep2 + i, dest_col_rep2 + 36).value = team2_over
        ws.cell(dest_row_rep2 + i, dest_col_rep2 + 37).value = team2_under
        team1_over, team1_under = count_over_under_offensive(matches, 9.5, 'sht', team1, last_matches_report2)
        team2_over, team2_under = count_over_under_offensive(matches, 9.5, 'sht', team2, last_matches_report2)
        ws.cell(dest_row_rep2 + i, dest_col_rep2 + 38).value = team1_over
        ws.cell(dest_row_rep2 + i, dest_col_rep2 + 39).value = team1_under
        ws.cell(dest_row_rep2 + i, dest_col_rep2 + 40).value = team2_over
        ws.cell(dest_row_rep2 + i, dest_col_rep2 + 41).value = team2_under
        # g_offensive
        team1_over, team1_under = count_over_under_offensive(matches, 0.5, 'g', team1, last_matches_report2)
        team2_over, team2_under = count_over_under_offensive(matches, 0.5, 'g', team2, last_matches_report2)
        ws.cell(dest_row_rep2 + i, dest_col_rep2 + 42).value = team1_over
        ws.cell(dest_row_rep2 + i, dest_col_rep2 + 43).value = team1_under
        ws.cell(dest_row_rep2 + i, dest_col_rep2 + 44).value = team2_over
        ws.cell(dest_row_rep2 + i, dest_col_rep2 + 45).value = team2_under
        team1_over, team1_under = count_over_under_offensive(matches, 1.5, 'g', team1, last_matches_report2)
        team2_over, team2_under = count_over_under_offensive(matches, 1.5, 'g', team2, last_matches_report2)
        ws.cell(dest_row_rep2 + i, dest_col_rep2 + 46).value = team1_over
        ws.cell(dest_row_rep2 + i, dest_col_rep2 + 47).value = team1_under
        ws.cell(dest_row_rep2 + i, dest_col_rep2 + 48).value = team2_over
        ws.cell(dest_row_rep2 + i, dest_col_rep2 + 49).value = team2_under
        team1_over, team1_under = count_over_under_offensive(matches, 2.5, 'g', team1, last_matches_report2)
        team2_over, team2_under = count_over_under_offensive(matches, 2.5, 'g', team2, last_matches_report2)
        ws.cell(dest_row_rep2 + i, dest_col_rep2 + 50).value = team1_over
        ws.cell(dest_row_rep2 + i, dest_col_rep2 + 51).value = team1_under
        ws.cell(dest_row_rep2 + i, dest_col_rep2 + 52).value = team2_over
        ws.cell(dest_row_rep2 + i, dest_col_rep2 + 53).value = team2_under
        # sht_defensive
        team1_over, team1_under = count_over_under_defensive(matches, 3.5, 'sht', team1, last_matches_report2)
        team2_over, team2_under = count_over_under_defensive(matches, 3.5, 'sht', team2, last_matches_report2)
        ws.cell(dest_row_rep2 + i, dest_col_rep2 + 54).value = team1_over
        ws.cell(dest_row_rep2 + i, dest_col_rep2 + 55).value = team1_under
        ws.cell(dest_row_rep2 + i, dest_col_rep2 + 56).value = team2_over
        ws.cell(dest_row_rep2 + i, dest_col_rep2 + 57).value = team2_under
        team1_over, team1_under = count_over_under_defensive(matches, 6.5, 'sht', team1, last_matches_report2)
        team2_over, team2_under = count_over_under_defensive(matches, 6.5, 'sht', team2, last_matches_report2)
        ws.cell(dest_row_rep2 + i, dest_col_rep2 + 58).value = team1_over
        ws.cell(dest_row_rep2 + i, dest_col_rep2 + 59).value = team1_under
        ws.cell(dest_row_rep2 + i, dest_col_rep2 + 60).value = team2_over
        ws.cell(dest_row_rep2 + i, dest_col_rep2 + 61).value = team2_under
        team1_over, team1_under = count_over_under_defensive(matches, 9.5, 'sht', team1, last_matches_report2)
        team2_over, team2_under = count_over_under_defensive(matches, 9.5, 'sht', team2, last_matches_report2)
        ws.cell(dest_row_rep2 + i, dest_col_rep2 + 62).value = team1_over
        ws.cell(dest_row_rep2 + i, dest_col_rep2 + 63).value = team1_under
        ws.cell(dest_row_rep2 + i, dest_col_rep2 + 64).value = team2_over
        ws.cell(dest_row_rep2 + i, dest_col_rep2 + 65).value = team2_under
        # g_defensive
        team1_over, team1_under = count_over_under_defensive(matches, 0.5, 'g', team1, last_matches_report2)
        team2_over, team2_under = count_over_under_defensive(matches, 0.5, 'g', team2, last_matches_report2)
        ws.cell(dest_row_rep2 + i, dest_col_rep2 + 66).value = team1_over
        ws.cell(dest_row_rep2 + i, dest_col_rep2 + 67).value = team1_under
        ws.cell(dest_row_rep2 + i, dest_col_rep2 + 68).value = team2_over
        ws.cell(dest_row_rep2 + i, dest_col_rep2 + 69).value = team2_under
        team1_over, team1_under = count_over_under_defensive(matches, 1.5, 'g', team1, last_matches_report2)
        team2_over, team2_under = count_over_under_defensive(matches, 1.5, 'g', team2, last_matches_report2)
        ws.cell(dest_row_rep2 + i, dest_col_rep2 + 70).value = team1_over
        ws.cell(dest_row_rep2 + i, dest_col_rep2 + 71).value = team1_under
        ws.cell(dest_row_rep2 + i, dest_col_rep2 + 72).value = team2_over
        ws.cell(dest_row_rep2 + i, dest_col_rep2 + 73).value = team2_under
        team1_over, team1_under = count_over_under_defensive(matches, 2.5, 'g', team1, last_matches_report2)
        team2_over, team2_under = count_over_under_defensive(matches, 2.5, 'g', team2, last_matches_report2)
        ws.cell(dest_row_rep2 + i, dest_col_rep2 + 74).value = team1_over
        ws.cell(dest_row_rep2 + i, dest_col_rep2 + 75).value = team1_under
        ws.cell(dest_row_rep2 + i, dest_col_rep2 + 76).value = team2_over
        ws.cell(dest_row_rep2 + i, dest_col_rep2 + 77).value = team2_under
        i += 1


def write_file(file, rep1_addr, rep2_addr, rep1_last, rep2_last):
    wb = load_workbook(filename=file)
    for sheet in sheets:
        df = pd.read_excel(file, sheet)
        rows_number = df.shape[0]
        ws = wb[sheet]
        matches = parse_main_data(ws, rows_number, cols_number)
        team_names = parse_team_names(ws, key_word='teams')
        last_matches = parse_last_matches(ws, key_word='home_team')
        add_report_1(ws, matches, team_names, rep1_addr, rep1_last)
        add_report_2(ws, matches, last_matches, rep2_addr, rep2_last)
    wb.save(filename=file)


if __name__ == "__main__":
    write_file('top_19.xlsx', 'AA3', 'BR47', 5, 16)
